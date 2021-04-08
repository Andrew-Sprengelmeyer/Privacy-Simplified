
# Modules used
from setup import *                     # Used for checking module installation
test_setup()                            # Runs the script for setup
from common_apk_module import *         # Used for easily finding app ID's
from google_play_scraper import app     # Used for Android App Store Scraper
from openpyxl import load_workbook      # Used for Reading Excel Spreadsheet cell values
from openpyxl.styles import Alignment   # Used for Aligning Excel Spreadsheet cell values
from tkinter import *                   # Used for GUI display

########################################################################################################
# Reading in the Database aka Excel Spreadsheet, 'Server_Backend_Database.xlsx'                        #
########################################################################################################

# Loads Workbook
wb = load_workbook("Resources\Database\Server_Backend_Database.xlsx")
ws = wb["Sheet1"]

# Lists used for database rows throughout progra
app_name = []
app_version = []
app_score = []
app_grade = []
app_popularity = []
app_type = []

# Reads in only App Name row from database
rows_iter = ws.iter_rows(min_col=1, min_row=3, max_col=1, max_row=ws.max_row)
app_name = [[cell.value for cell in row] for row in rows_iter]

# Reads in only App Version row from database
rows_iter = ws.iter_rows(min_col=2, min_row=3, max_col=2, max_row=ws.max_row)
app_version = [[cell.value for cell in row] for row in rows_iter]

# Reads in only App Score row from database
rows_iter = ws.iter_rows(min_col=3, min_row=3, max_col=3, max_row=ws.max_row)
app_score = [[cell.value for cell in row] for row in rows_iter]

# Reads in only App Grade row from database
rows_iter = ws.iter_rows(min_col=4, min_row=3, max_col=4, max_row=ws.max_row)
app_grade = [[cell.value for cell in row] for row in rows_iter]

# Reads in only App Popularity row from database
rows_iter = ws.iter_rows(min_col=5, min_row=3, max_col=5, max_row=ws.max_row)
app_popularity = [[cell.value for cell in row] for row in rows_iter]

# Reads in only App Type row from database
rows_iter = ws.iter_rows(min_col=6, min_row=3, max_col=6, max_row=ws.max_row)
app_type = [[cell.value for cell in row] for row in rows_iter]

# Convert App Name list of a list to just a regular list
app_name = [item for sublist in app_name for item in sublist]

# Convert App Version list of a list to just a regular list
app_version = [item for sublist in app_version for item in sublist]

# Convert App Score list of a list to just a regular list
app_score = [item for sublist in app_score for item in sublist]

# Convert App Grade list of a list to just a regular list
app_grade = [item for sublist in app_grade for item in sublist]

# Convert App Popular list of a list to just a regular list
app_popularity = [item for sublist in app_popularity for item in sublist]

# Convert App Type list of a list to just a regular list
app_type = [item for sublist in app_type for item in sublist]

########################################################################################################
# GUI Setup                                                                                            #
########################################################################################################

root = Tk()
root.title('CI475: Cyber Senior Thesis')
root.geometry("400x650")

# make sure app cannot be resized
root.resizable(width=False, height=False)

# Define image
bg = PhotoImage(file="Resources\pictures\iPhone_and_Background.png")

# Create a canvas
my_canvas = Canvas(root, width=400, height=650)
my_canvas.pack(fill="both", expand=True)

# Set image in canvas
my_canvas.create_image(0,20, image=bg, anchor="nw")

# Add a label
my_canvas.create_text(210, 75, text="Privacy Simplified", font=("Times", 26), fill="white")
my_canvas.create_text(210, 105, text="Privacy made simple", font=("Times", 14, "italic"), fill="white")

# Search Bar
app_name_entry = Entry(root, font=("Times", 18), width=18, fg="#336d92", bd=0)
app_name_entry_window = my_canvas.create_window(100, 200, anchor="nw", window=app_name_entry)

# Filler Text in Search Bar
app_name_entry.insert(END, "Search for an app") 

########################################################################################################
# Dictionary to reference                                                                              #
########################################################################################################

# Dictionary of Criteria for Privacy Impact assessment (PIA)
main_app_criteria_and_points_dictionary = {}

with open("Resources\miscellaneous_files\main_app_criteria_and_points.txt","r") as text:
    for line in text:
        key, value = line.split(",")
        if int(value) > main_app_criteria_and_points_dictionary.get(key, -1):
            main_app_criteria_and_points_dictionary[key] = int(value)

########################################################################################################
# GUI Application Module for Deleting Landing Page Canvas                                              #
########################################################################################################

def delete_landing_page_canvas_function():
    
    global search_app_name_window, popular_apps_button_window, settings_button_window, privacy_button_window, about_us_button_window
    
    my_canvas.delete(search_app_name_window)
    my_canvas.delete(popular_apps_button_window)
    my_canvas.delete(settings_button_window)
    my_canvas.delete(privacy_button_window)
    my_canvas.delete(about_us_button_window)

########################################################################################################
# Converts app score to app grade                                                                      #
########################################################################################################

def app_grade_function(app_score):
    if (app_score >= 421): 
        app_grade = "F"
    elif (app_score <= 420):
        app_grade = "D"
    elif (app_score <= 315):
        app_grade = "C"
    elif (app_score <= 210):
        app_grade = "B"
    elif (app_score <= 105):
        app_grade = "A"
    else:
        app_grade = "?"
    return app_grade

########################################################################################################
# Displays additional details i.e. criteria and points                                                 #
########################################################################################################

def view_details_page(): 
    view_details_page_button_window = 14
    my_canvas.delete(view_details_page_button_window)
    
    global listbox_criteria, listbox_points, rectangle_filled_details, criteria, points

    delete_landing_page_canvas_function()
    
    # PIA criteria
    app_criteria_pia_list = list(main_app_criteria_and_points_dictionary.keys())
    
    # Removes zeroth index/first value ("App Criteria")
    del app_criteria_pia_list[0]

    # Points awarded for PIA criteria in a list
    app_criteria_points_list = list(main_app_criteria_and_points_dictionary.values())

    # Removes zeroth index/first value ("Points Worth")
    del app_criteria_points_list[0]
    
    # Rectangle Background
    rectangle_filled_details = my_canvas.create_rectangle(100, 250, 320, 475, outline="gray1", fill="#C7E6E4")
    
    #############################################################################################
    # miscellaneous Functions for Listboxes                                                     #
    #############################################################################################
    
    def scrolllistbox2(event):
        listbox_points.yview_scroll(int(-4*(event.delta/120)), "units")
    def scrolllistbox1(event):
        listbox_criteria.yview_scroll(int(-4*(event.delta/120)), "units")
    
    #############################################################################################
    # App Criteria Listbox                                                                      #
    #############################################################################################
    
    criteria = my_canvas.create_text(150, 270, text="Criteria", font=("Times", 14, "bold"), fill="black")

    scrollbar = Scrollbar(root)
    scrollbar.pack(side=RIGHT, fill=Y)
    listbox_criteria = Listbox(root, height=11, width=20)
    listbox_criteria.place(x = 115, y = 285)
    
    for i in app_criteria_pia_list:
        listbox_criteria.insert("end", i)
        
    # attach listbox to scrollbar
    listbox_criteria.config(yscrollcommand=scrollbar.set)
    listbox_criteria.bind("<MouseWheel>", scrolllistbox2)
    
    #############################################################################################
    # App Points Listbox                                                                        #
    #############################################################################################
    
    points = my_canvas.create_text(270, 270, text="Points", font=("Times", 14, "bold"), fill="black")

    listbox_points = Listbox(root, height=11, width=2) 
    listbox_points.place(x = 270, y = 285)
    
    for i in app_criteria_points_list:
        listbox_points.insert("end", i)
        
    # attach listbox to scrollbar
    listbox_points.config(yscrollcommand=scrollbar.set)
    listbox_points.bind("<MouseWheel>", scrolllistbox1)

########################################################################################################
# Displays app grade on GUI                                                                            #
########################################################################################################

def display_app_result_page(app_grade):

    global rectangle_filled_details, view_details_page_button_window, app_grade_canvas

    delete_landing_page_canvas_function()

    rectangle_filled_details = my_canvas.create_rectangle(100, 250, 320, 475, outline="gray1", fill="#C7E6E4")
    
    view_details_page_button = Button(root, text="View Details", command=view_details_page)
    view_details_page_button_window = my_canvas.create_window(175, 440, anchor="nw", window=view_details_page_button)

    if (app_grade == "F"):
        app_grade_canvas = my_canvas.create_text(220, 360, text="F", font=("Times", 100), fill="red2")
        return app_grade_canvas
    elif (app_grade == "D"):
        app_grade_canvas = my_canvas.create_text(220, 360, text="D", font=("Times", 100), fill="DarkOrange2")
        return app_grade_canvas
    elif (app_grade == "C"):
        app_grade_canvas = my_canvas.create_text(220, 360, text="C", font=("Times", 100), fill="DarkGoldenrod2")
        return app_grade_canvas
    elif (app_grade == "B"):
        app_grade_canvas = my_canvas.create_text(220, 360, text="B", font=("Times", 100), fill="yellow2")
        return app_grade_canvas
    elif (app_grade == "A"):
        app_grade_canvas = my_canvas.create_text(220, 360, text="A", font=("Times", 100), fill="green2")
        return app_grade_canvas
    else:
        app_grade_canvas = my_canvas.create_text(220, 360, text="?", font=("Times", 100), fill="blue2")
        return app_grade_canvas

########################################################################################################
# Displays App Score on Screen                                                                         #
########################################################################################################

def display_app_score(app_score_write):
    
    global canvas_filled_app_score_text, canvas_filled_app_score
    
    delete_landing_page_canvas_function()
    
    # NOT NECCESARY ANYMORE!!!
    # Normalize the app score
    # It is out of 1110 initially then, it gets put into a value from 0 - 114
    # app_score_write /= 10
    # app_score_write -= 100
    # app_score_write = abs(app_score_write)

    # App Score printout
    canvas_filled_app_score_text = my_canvas.create_text(210, 270, text="App Score:    ", font=("Times", 14, "bold"), fill="black")
    canvas_filled_app_score = my_canvas.create_text(270, 270, text=app_score_write, font=("Times", 14), fill="black")

########################################################################################################
# GUI Application Module for PIA                                                                       #
# Calculates Privacy App Score                                                                         #
########################################################################################################

def app_score_privacy_impact_assessment(app_name_write):

    # Makes file name findable
    file_name = app_name_write + ".apk"
    
    file_path = "Resources\\app_store\\"
    
    full_file_name = file_path + file_name
    
    app_score_write = 0
        
    # PIA criteria
    app_criteria_pia_list = list(main_app_criteria_and_points_dictionary.keys())
    
    # Removes zeroth index/first value ("App Criteria")
    del app_criteria_pia_list[0]
    
    # Points awarded for PIA criteria in a list
    app_criteria_points_list = list(main_app_criteria_and_points_dictionary.values())

    # Removes zeroth index/first value ("Points Worth")
    del app_criteria_points_list[0]

    # Builds the app criteria list from read in application
    with open(full_file_name) as file:
        app_criteria_list = file.readlines()
    
    file.close()
    
    # Strips new line from app criteria list
    app_criteria_list = [i[:-1] for i in app_criteria_list]

    for index, (first, second) in enumerate(zip(app_criteria_pia_list, app_criteria_list)):
        if first != second:
            app_score_write += app_criteria_points_list[index]
        elif (first == second):
            app_score_write += 0

    # NOT NECCESARY ANYMORE!!!
    # It is out of 1110 initially then, it gets put into a value from 0 - 114
    # app_score_write /= 10

    return app_score_write

########################################################################################################
# GUI Application Module for Search                                                                    #
# Allows for search bar to function on GUI                                                             #  
########################################################################################################

def search():
    
    global app_result_window
    
    delete_landing_page_canvas_function()
    
    app_result_window = my_canvas.create_text(220, 178, text="App Result", font=("Times", 14), fill="white")
        
    # Gathering user input from search bar
    app_name_query = app_name_entry.get()
    
    # Convert user input to app id
    app_name_query_to_app_id = return_app_id(app_name_query)
    
    # using app id to get google play store scraped information returns a dictionary page of app page 
    result = app(app_name_query_to_app_id, lang="en", country="us")
    
    # Searching for version ID in result dictionary    
    if ("version" in result):
        app_version_query = result["version"]

    # Searching for genre in result dictionary    
    if ( "genre" in result):
        app_genre = result["genre"]

    # First level if (searched app name is not on list)
    if (app_name_query not in app_name):
        
        # Appends App Name to their respective list
        app_name.append(app_name_query)

        # Appends App version to their respective list
        app_version.append(app_version_query)

        # Appends App Genre to their respective list
        app_type.append(app_genre)
        
        # Increases row max to account for new addition
        row_max = ws.max_row + 1

        # Write only App Name
        app_name_write = (str(app_name[-1]))
        ws.cell(row=row_max, column=1).value=app_name_write
        wb.save("Resources\Database\Server_Backend_Database.xlsx")

        # Write only App Versions
        app_version_write = str(app_version[-1])
        ws.cell(row=row_max, column=2).value=app_version_write # App Version ID Number may vary by phone
        wb.save("Resources\Database\Server_Backend_Database.xlsx")

        app_score_write = app_score_privacy_impact_assessment(app_name_write)

        # Write only App Grades
        app_grade_write = app_grade_function(app_score_write)
        ws.cell(row=row_max, column=4).value=app_grade_write
        wb.save("Resources\Database\Server_Backend_Database.xlsx")
        display_app_result_page(app_grade_write)
        
        # Write only App Scores
        app_score_write = app_score_privacy_impact_assessment(app_name_write)
        ws.cell(row=row_max, column=3).value=app_score_write
        wb.save("Resources\Database\Server_Backend_Database.xlsx")
        display_app_score(app_score_write)

        ########################################################################################################
        # POPULARITY MARKING                                                                                   #
        ########################################################################################################

        # Write only Popularity Tally
        # Find App Index
        app_index = app_name.index(app_name_query)
    
        # Creation of App Popularity Value
        app_popularity_value = 0
        app_popular_addition = 1        
        
        app_popularity_value = app_popular_addition + app_popularity_value
        
        app_popularity.insert(app_index, app_popularity_value)
        
        ws.cell(row=row_max, column=5).value=app_popularity_value
        wb.save("Resources\Database\Server_Backend_Database.xlsx")
        
        # Write only App Type
        app_genre_write = (str(app_type[-1]))
        ws.cell(row=row_max, column=6).value=app_genre_write
        wb.save("Resources\Database\Server_Backend_Database.xlsx")
        
    # First level else (searched app name is on list)
    else:
        # Second layer if (searched app version is not on list, but app name is on list)
        if (app_version_query not in app_version):

            # Appends App Name to their respective list
            app_name.append(app_name_query)

            # Appends App version to their respective list
            app_version.append(app_version_query)

            # Appends App Genre to their respective list
            app_type.append(app_genre)
            
            # Increases row max to account for new addition
            row_max = ws.max_row + 1

            # Write only App Name
            app_name_write = (str(app_name[-1]))
            ws.cell(row=row_max, column=1).value=app_name_write
            wb.save("Resources\Database\Server_Backend_Database.xlsx")

            # Write only App Versions
            app_version_write = str(app_version[-1])
            ws.cell(row=row_max, column=2).value=app_version_write
            wb.save("Resources\Database\Server_Backend_Database.xlsx")

            app_score_write = app_score_privacy_impact_assessment(app_name_write)

            # Write only App Grades
            app_grade_write = app_grade_function(app_score_write)
            ws.cell(row=row_max, column=4).value=app_grade_write
            wb.save("Resources\Database\Server_Backend_Database.xlsx")
            display_app_result_page(app_grade_write)
            
            # Write only App Scores                                                                                                              
            app_score_write = app_score_privacy_impact_assessment(app_name_write)
            ws.cell(row=row_max, column=3).value=app_score_write
            wb.save("Resources\Database\Server_Backend_Database.xlsx")
            display_app_score(app_score_write)
            
            ########################################################################################################
            # POPULARITY MARKING                                                                                   #
            ########################################################################################################
            
            # Write only Popularity Tally
            # Find App Index
            app_index = app_name.index(app_name_query)
            
            # Creation of App Popularity Value
            app_popularity_value = 0
            app_popular_addition = 1        
            
            app_popularity_value = app_popular_addition + app_popularity_value
            
            app_popularity.insert(app_index, app_popularity_value)
            
            ws.cell(row=row_max, column=5).value=app_popularity_value
            wb.save("Resources\Database\Server_Backend_Database.xlsx")
            
            # Write only App Type
            app_genre_write = (str(app_type[-1]))
            ws.cell(row=row_max, column=6).value=app_genre_write
            wb.save("Resources\Database\Server_Backend_Database.xlsx")
        
        # Second layer else (searched app version is on list and app name is on list)
        else:
            app_index = app_name.index(app_name_query)
            app_grade_write = app_grade[app_index]
            app_score_write = app_score[app_index]
            app_genre_write = app_type[app_index]
            display_app_result_page(app_grade_write)
            display_app_score(app_score_write)
            
            ########################################################################################################
            # POPULARITY MARKING                                                                                   #
            ########################################################################################################
            
            # Write only Popularity Tally
            # Find App Index
            app_index = app_name.index(app_name_query)
    
            # Creation of App Popularity Value
            app_popularity_value = 0
            app_popular_addition = 1
            
            app_popularity_value = app_popularity[app_index]
            
            app_popularity_value = app_popular_addition + app_popularity_value
            
            app_popularity.insert(app_index, app_popularity_value)
            
            row_number = app_index + 3 # To account for the database header and table headings and +1 shift from 0 to 1
            
            ws.cell(row=row_number, column=5).value=app_popularity_value
            wb.save("Resources\Database\Server_Backend_Database.xlsx")
    
########################################################################################################
# GUI Application Module for Popular Apps                                                              #
########################################################################################################

def popular_apps_page():

    global app_name_window, popular_apps_window, rectangle_filled_popular_apps, app_grade_window, listbox_name, listbox_grade

    delete_landing_page_canvas_function()

    updated_app_popularity_index = []
    popular_app_name_list = []   
    popular_app_grade_list = []
    
    # Makes the list go in order from highest to lowest values
    updated_app_popularity_index = sorted( [(x,i) for (i,x) in enumerate(app_popularity)], reverse=True )[:(len(app_popularity))]

    # Reduces down to just one list instead of sorted list
    updated_app_popularity_index = [item for i in updated_app_popularity_index for item in i]
    
    # Gets rid of every other value which was a 0
    updated_app_popularity_index = updated_app_popularity_index[1::2]
       
    # Re Organizes list based off of the indexes provided by updated_app_popularity_index
    for i in updated_app_popularity_index:
        popular_app_name_list.insert(i, app_name[i])
    
    # Re Organizes list based off of the indexes provided by updated_app_popularity_index
    for i in updated_app_popularity_index:
        popular_app_grade_list.insert(i, app_grade[i])

    popular_apps_window = my_canvas.create_text(220, 178, text="Popular Apps", font=("Times", 14), fill="white")    
    rectangle_filled_popular_apps = my_canvas.create_rectangle(100, 250, 320, 475, outline="gray1", fill="#C7E6E4")
    
    #############################################################################################
    # miscellaneous Functions for Listboxes                                                     #
    #############################################################################################
    
    def scrolllistbox2(event):
        listbox_grade.yview_scroll(int(-4*(event.delta/120)), "units")
    def scrolllistbox1(event):
        listbox_name.yview_scroll(int(-4*(event.delta/120)), "units")
    
    #############################################################################################
    # App Name Listbox                                                                          #
    #############################################################################################
    
    app_name_window = my_canvas.create_text(150, 270, text="App Name", font=("Times", 14, "bold"), fill="black")

    scrollbar = Scrollbar(root)
    scrollbar.pack(side=RIGHT, fill=Y)
    listbox_name = Listbox(root, height=11, width=17)
    listbox_name.place(x = 105, y = 285)
    
    for i in popular_app_name_list:
        listbox_name.insert("end", i)
        
    # attach listbox to scrollbar
    listbox_name.config(yscrollcommand=scrollbar.set)
    listbox_name.bind("<MouseWheel>", scrolllistbox2)
    
    #############################################################################################
    # App Grade Listbox                                                                         #
    #############################################################################################
    
    app_grade_window = my_canvas.create_text(270, 270, text="App Grade", font=("Times", 14, "bold"), fill="black")

    listbox_grade = Listbox(root, height=11, width=3) 
    listbox_grade.place(x = 250, y = 285)
    
    for i in popular_app_grade_list:
        listbox_grade.insert("end", i)
        
    # attach listbox to scrollbar
    listbox_grade.config(yscrollcommand=scrollbar.set)
    listbox_grade.bind("<MouseWheel>", scrolllistbox1)
        
########################################################################################################
# GUI Application Module for Notifications                                                             #
########################################################################################################

def notifications_page(): 

    global notifications_window

    delete_landing_page_canvas_function()
    # delete_settings_page_canvas_function()
    
    notifications_window = my_canvas.create_text(220, 178, text="Notifications", font=("Times", 14), fill="white")

    print("Notifications Page") 
    
########################################################################################################
# GUI Application Module for Appearence                                                                #
########################################################################################################

def appearence_page(): 

    global appearence_window

    delete_landing_page_canvas_function()
    # delete_settings_page_canvas_function()
    
    appearence_window = my_canvas.create_text(220, 178, text="Appearence", font=("Times", 14), fill="white")

    print("Appearence Page")
   
########################################################################################################
# GUI Application Module for Settings                                                                  #
########################################################################################################

def settings_page(): 

    global settings_window, notifications_button_window, appearence_button_window, privacy_button_window

    delete_landing_page_canvas_function()
    # delete_privacy_page_canvas_function()
    
    settings_window = my_canvas.create_text(220, 178, text="Settings", font=("Times", 14), fill="white")
                                                                                                                            
    notifications_button = Button(root, bg="#C7E6E4", text="Notifications", command=notifications_page)                    
    appearence_button = Button(root, bg="#C7E6E4", text="Appearence", command=appearence_page)                              
    privacy_button = Button(root, bg="#C7E6E4", text="Privacy", command=privacy_page)                                         

    notifications_button_window = my_canvas.create_window(180, 240, anchor="nw", window=notifications_button)
    appearence_button_window = my_canvas.create_window(180, 280, anchor="nw", window=appearence_button)
    privacy_button_window = my_canvas.create_window(180, 320, anchor="nw", window=privacy_button)
    
########################################################################################################
# GUI Application Module for Privacy                                                                   #
########################################################################################################

def privacy_page():

    global privacy_window, privacy_frame
   
    delete_landing_page_canvas_function()
    # delete_settings_page_canvas_function()

    privacy_window = my_canvas.create_text(220, 178, text="Privacy", font=("Times", 14), fill="white")
    root.configure(background="#808000")
    privacy_frame = Frame(root, width=80, height=80, bg = '#ffffff', borderwidth=3, relief="sunken")
    scrollbar = Scrollbar(privacy_frame) 
    edit_area = Text(privacy_frame, width=20, height=12, wrap="word", yscrollcommand=scrollbar.set, borderwidth=3, highlightthickness=3)
    scrollbar.config(command=edit_area.yview)
    scrollbar.pack(side="right", fill="y")
    edit_area.pack(side="left", fill="both", expand=True)

    with open("Resources\important_documents\privacy_policy.txt", "r") as file:
        data = file.read()
    file.close()
    
    edit_area.insert(INSERT, data)
    privacy_frame.place(x=115,y=250)
    
########################################################################################################
# GUI Application Module for About Us                                                                  #
########################################################################################################

def about_us_page():
    
    global about_us_window, about_us_frame

    delete_landing_page_canvas_function()

    about_us_window = my_canvas.create_text(220, 178, text="About Us", font=("Times", 14), fill="white")

    # Based off of https://stackoverflow.com/questions/7727804/tkinter-using-scrollbars-on-a-canvas

    root.configure(background="#808000")
    about_us_frame = Frame(root, width=80, height=80, bg = '#ffffff', borderwidth=3, relief="sunken")
    scrollbar = Scrollbar(about_us_frame) 
    edit_area = Text(about_us_frame, width=20, height=12, wrap="word", yscrollcommand=scrollbar.set, borderwidth=3, highlightthickness=3)
    scrollbar.config(command=edit_area.yview)
    scrollbar.pack(side="right", fill="y")
    edit_area.pack(side="left", fill="both", expand=True)
        
    edit_area.insert(INSERT, """This is the Python GUI proof of concept for Privacy Simplified for Andrew Sprengelmeyer's CI475 Thesis""")
    about_us_frame.place(x=115,y=250)
    
########################################################################################################
# GUI Application Module for Landing Page                                                              #
########################################################################################################
    
def landing_page():

    global search_app_name_window, popular_apps_button_window, settings_button_window, privacy_button_window, about_us_button_window

    search_app_name_window = my_canvas.create_text(220, 178, text="Search App Name", font=("Times", 14), fill="white")

    # add some buttons
    popular_apps_button = Button(root, bg="#C7E6E4", text="Popular Apps", command=popular_apps_page)
    settings_button = Button(root, bg="#C7E6E4", text="Settings", command=settings_page)
    privacy_button = Button(root, bg="#C7E6E4", text="Privacy", command=privacy_page)
    about_us_button = Button(root, bg="#C7E6E4", text="About us", command=about_us_page)

    popular_apps_button_window = my_canvas.create_window(180, 240, anchor="nw", window=popular_apps_button)
    settings_button_window = my_canvas.create_window(180, 280, anchor="nw", window=settings_button)
    privacy_button_window = my_canvas.create_window(180, 320, anchor="nw", window=privacy_button)
    about_us_button_window = my_canvas.create_window(180, 360, anchor="nw", window=about_us_button)

# Starts the Program
landing_page()

# Search button and clicking function
search_button_image = PhotoImage(file="Resources\pictures\search.png")
search_button_image = search_button_image.subsample(43,43) # Downsizes to correct better size
search_button = Button(root, text="search", image=search_button_image, command=search)
search_button_window = my_canvas.create_window(290, 200, anchor="nw", window=search_button)

# Home button
home_button_image = PhotoImage(file="Resources\pictures\home.png")
home_button_image = home_button_image.subsample(2,2)
home_button = Button(root, text="home", image=home_button_image, command=landing_page)
home_button_window = my_canvas.create_window(180, 500, anchor="nw", window=home_button)

# Sets up infinite loop for GUI
root.mainloop()
